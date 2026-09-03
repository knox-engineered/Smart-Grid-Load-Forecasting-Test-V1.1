"""
Extract regional 'Energy Met (MU)' from the MOP_E sheet of each daily NLDC
PSP .xls file, and compute month-averaged regional demand-share weights.

Usage:
    python3 extract_regional_weights.py /path/to/uploads/*.xls
"""
import sys, re, subprocess, os
import pandas as pd
from openpyxl import load_workbook

SOFFICE_WRAPPER = "/mnt/skills/public/xlsx/scripts/office/soffice.py"
WORKDIR = "./converted"
os.makedirs(WORKDIR, exist_ok=True)

REGIONS = ["NR", "WR", "SR", "ER", "NER"]


def date_from_filename(fname):
    m = re.match(r"(\d{2})_(\d{2})_(\d{2})", os.path.basename(fname))
    if not m:
        return None
    dd, mm, yy = m.groups()
    return pd.to_datetime(f"20{yy}-{mm}-{dd}")


def convert_xls_to_xlsx(xls_path):
    base = os.path.splitext(os.path.basename(xls_path))[0]
    out_path = os.path.join(WORKDIR, base + ".xlsx")
    if os.path.exists(out_path):
        return out_path
    subprocess.run(
        ["python3", SOFFICE_WRAPPER, "--headless", "--convert-to", "xlsx",
         "--outdir", WORKDIR, xls_path],
        check=True, capture_output=True
    )
    return out_path


def extract_energy_met(xlsx_path):
    """Find the 'Energy Met (MU)' row in MOP_E sheet and return {region: value}."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["MOP_E"]
    rows = list(ws.iter_rows(values_only=True))

    header_row_idx = None
    for i, r in enumerate(rows):
        if r[2] == "NR" and r[3] == "WR":
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError("Could not find region header row (NR, WR, SR...)")

    for r in rows[header_row_idx + 1:]:
        label = r[0]
        if isinstance(label, str) and "Energy Met" in label and "Shortage" not in label:
            values = r[2:7]  # NR, WR, SR, ER, NER columns
            return dict(zip(REGIONS, values))

    raise ValueError("Could not find 'Energy Met (MU)' row")


def main(files):
    rows = []
    for f in files:
        day_date = date_from_filename(f)
        if day_date is None:
            print(f"Skipping (bad filename pattern): {f}")
            continue
        xlsx_path = convert_xls_to_xlsx(f)
        try:
            energy = extract_energy_met(xlsx_path)
            energy["date"] = day_date
            rows.append(energy)
            print(f"OK: {f} -> {day_date.date()} : {energy}")
        except Exception as e:
            print(f"FAILED: {f} -> {e}")

    if not rows:
        print("No data extracted.")
        return

    df = pd.DataFrame(rows).set_index("date").sort_index()
    df.to_csv("regional_energy_met_daily.csv")

    # Compute daily shares, then average across the month
    daily_totals = df[REGIONS].sum(axis=1)
    daily_shares = df[REGIONS].div(daily_totals, axis=0)

    month_avg_weights = daily_shares.mean()
    month_avg_weights = month_avg_weights / month_avg_weights.sum()  # renormalize to exactly 1.0

    print("\n=== Daily regional Energy Met (MU) ===")
    print(df)
    print("\n=== Month-averaged regional demand-share weights ===")
    print(month_avg_weights)

    month_avg_weights.to_csv("regional_weights.csv", header=["weight"])
    print("\nSaved: regional_energy_met_daily.csv, regional_weights.csv")


if __name__ == "__main__":
    main(sys.argv[1:])
